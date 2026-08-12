#!/usr/bin/env python3
"""Parse Jameel supplier/customer August ledgers and emit data/reports artifacts.
Same column rules as petroleum_data_import.import_wizard.
"""
from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

GRADES = ("PMS", "AGO", "IK")
CONTROL_NAMES = {
    "GROSS JAMEEL", "NET JAMEEL", "JAMEEL CUSTOMERS", "JAMEEL SUPPLIERS",
    "JAMEEL CUSTOMER", "JAMEEL SUPPLIER",
}


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s)
        except ValueError:
            pass
        for sep in ("/", "-", "."):
            if sep in s:
                parts = s.split(sep)
                if len(parts) == 3:
                    try:
                        a, b, c = (int(p) for p in parts)
                    except ValueError:
                        return None
                    if a > 31:
                        y, m, d = a, b, c
                    else:
                        d, m, y = a, b, c
                    if y < 100:
                        y += 2000
                    try:
                        return date(y, m, d)
                    except ValueError:
                        return None
    return None


def _is_header(cells):
    return "BALANCE" in cells and ("DEBIT" in cells or "CREDIT" in cells)


def _build_colmap(header):
    cm, grade = {}, None
    for idx, h in enumerate(header):
        if h in ("LOADING", "LOADING DATE", "DATE"):
            cm.setdefault("date", idx)
        elif h == "LOADING POINT":
            cm["lp"] = idx
        elif h in ("TRUCKS", "TRUCK"):
            cm["truck"] = idx
        elif h in GRADES:
            cm[h.lower()] = idx
            grade = h.lower()
        elif h == "PRICE" and grade:
            cm[grade + "_price"] = idx
            grade = None
        elif h == "SELLING PRICE":
            cm["sp"] = idx
        elif h == "INVOICE NO":
            cm["inv"] = idx
        elif h == "TRANS DATE":
            cm["tdate"] = idx
        elif h == "DEBIT":
            cm["debit"] = idx
        elif h == "CREDIT":
            cm["credit"] = idx
        elif h == "BALANCE":
            cm["balance"] = idx
    return cm


def parse_sheet(ws, default_side):
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=18, values_only=True))
    up = [[str(c).strip().upper() if c is not None else "" for c in r] for r in rows]
    header_idxs = [i for i, c in enumerate(up) if _is_header(c)]
    sections = []
    for si, hi in enumerate(header_idxs):
        end = header_idxs[si + 1] if si + 1 < len(header_idxs) else len(rows)
        cm = _build_colmap(up[hi])
        if "balance" not in cm:
            continue
        side = default_side
        for j in range(max(0, hi - 3), hi + 1):
            blob = " ".join(up[j])
            if "PAYABLE" in blob:
                side = "ap"
            elif "RECEIVABLE" in blob:
                side = "ar"
        bf, txns, last_balance = 0.0, [], None
        for r in rows[hi + 1 : end]:
            def get(k, r=r, cm=cm):
                i = cm.get(k)
                return r[i] if (i is not None and i < len(r)) else None

            bal = get("balance")
            if isinstance(bal, (int, float)):
                last_balance = bal
            sp = get("sp")
            if isinstance(sp, str) and sp.strip().upper() == "B/F":
                bf = _num(bal)
                continue
            d = _parse_date(get("date")) or _parse_date(get("tdate"))
            if d is None:
                continue
            debit, credit = _num(get("debit")), _num(get("credit"))
            if debit == 0 and credit == 0:
                continue
            lp = str(get("lp")).strip().upper() if get("lp") else ""
            effect = (debit - credit) if side == "ar" else (credit - debit)
            lines = []
            for g in ("pms", "ago", "ik"):
                q, p = _num(get(g)), _num(get(g + "_price"))
                if q and p:
                    lines.append({"grade": g.upper(), "qty": q, "price": p})
            if lp in ("PAYMENT", "REFUND") or (not lines and _num(sp) == 0):
                kind = "refund" if lp == "REFUND" else "payment"
            else:
                kind = "loading"
            txns.append({
                "date": d.isoformat(), "effect": effect, "debit": debit, "credit": credit,
                "kind": kind, "lp": lp, "lines": lines, "sp": _num(sp),
                "truck": str(get("truck") or "").strip().upper(),
                "inv": str(get("inv")) if get("inv") not in (None, "") else "",
            })
        sections.append({"side": side, "bf": bf, "txns": txns, "last_balance": last_balance or 0.0})
    return sections


def partner_name(ws, sn):
    title = None
    for r in ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True):
        title = r[0]
    name = (str(title).strip() if title else sn).strip()
    for suf in ("(PAYABLE)", "(RECEIVABLE)"):
        name = name.replace(suf, "").replace(suf.lower(), "").strip()
    return name or sn


def load_workbook_txns(path, default_side):
    wb = load_workbook(path, data_only=True)
    partners = []
    try:
        for sn in wb.sheetnames:
            if sn in ("Sheet1", "Sheet2"):
                continue
            ws = wb[sn]
            name = partner_name(ws, sn)
            if name.upper() in CONTROL_NAMES or sn.upper() in CONTROL_NAMES:
                continue
            secs = parse_sheet(ws, default_side)
            if secs:
                partners.append({"sheet": sn, "name": name, "sections": secs})
    finally:
        wb.close()
    return partners


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--suppliers", required=True)
    ap.add_argument("--customers", required=True)
    ap.add_argument("--from", dest="d0", default="2026-08-01")
    ap.add_argument("--to", dest="d1", default="2026-08-31")
    ap.add_argument("--out", default="data/reports")
    args = ap.parse_args()
    d0, d1 = date.fromisoformat(args.d0), date.fromisoformat(args.d1)
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    suppliers = load_workbook_txns(args.suppliers, "ap")
    customers = load_workbook_txns(args.customers, "ar")

    def flatten(partners, role):
        rows = []
        for p in partners:
            for sec in p["sections"]:
                for t in sec["txns"]:
                    d = date.fromisoformat(t["date"])
                    if d0 <= d <= d1:
                        rows.append({"role": role, "partner": p["name"], "sheet": p["sheet"], "side": sec["side"], **t})
        return rows

    all_rows = flatten(suppliers, "supplier") + flatten(customers, "customer")
    with open(out / "aug_all_transactions.csv", "w", newline="") as f:
        fields = ["role", "partner", "sheet", "side", "date", "kind", "lp", "truck", "inv", "debit", "credit", "effect", "sp", "lines"]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in sorted(all_rows, key=lambda x: (x["date"], x["role"], x["partner"])):
            row = {k: r.get(k, "") for k in fields}
            row["lines"] = json.dumps(r["lines"])
            w.writerow(row)
    print(f"Wrote {len(all_rows)} txns to {out / 'aug_all_transactions.csv'}")


if __name__ == "__main__":
    main()
