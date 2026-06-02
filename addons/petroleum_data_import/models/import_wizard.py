import base64
import datetime
import io
import logging

from odoo import Command, _, fields, models
from odoo.exceptions import UserError

from odoo.addons.petroleum_data_import.models.data_import_job import IMPORT_CTX, SECTION_BATCH

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

CONTROL_NAMES = {
    'GROSS JAMEEL', 'NET JAMEEL', 'JAMEEL CUSTOMERS', 'JAMEEL SUPPLIERS',
    'JAMEEL CUSTOMER', 'JAMEEL SUPPLIER',
}
GRADES = ('PMS', 'AGO', 'IK')

# Payment "payer" names that are really JAMEEL's own bank/cash accounts.
# name on the worksheet -> (journal name, journal code)
BANK_NAME_MAP = {
    'ABSA': ('ABSA Bank', 'ABSA'),
    'KCB': ('KCB Bank', 'KCB'),
    'EQUITY': ('Equity Bank', 'EQTY'),
    'PREMIER': ('Premier Bank', 'PREM'),
    'GULF': ('Gulf African Bank', 'GULF'),
    'MPESA': ('M-Pesa', 'MPSA'),
}


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def _parse_date(v):
    """Accept real datetimes and text dates like '30/4/2026' or '2026-05-06'."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return datetime.date.fromisoformat(s)
        except ValueError:
            pass
        for sep in ('/', '-', '.'):
            if sep in s:
                parts = s.split(sep)
                if len(parts) == 3:
                    try:
                        a, b, c = (int(p) for p in parts)
                    except ValueError:
                        return None
                    if a > 31:
                        y, m, d = a, b, c
                    else:
                        d, m, y = a, b, c
                    if y < 100:
                        y += 2000
                    try:
                        return datetime.date(y, m, d)
                    except ValueError:
                        return None
    return None


class PetroleumDataImport(models.TransientModel):
    _name = 'petroleum.data.import'
    _description = 'Petroleum Excel Ledger Importer'

    company_id = fields.Many2one(
        'res.company', string='Company', required=True,
        default=lambda self: self.env.company)
    cutoff_date = fields.Date(
        string='Cut-over Date', required=True, default='2026-05-18',
        help='Opening balances are posted the day before this date; every '
             'transaction dated on/after it is imported individually.')

    customers_file = fields.Binary(string='Customers Workbook (.xlsx)')
    customers_filename = fields.Char()
    suppliers_file = fields.Binary(string='Suppliers Workbook (.xlsx)')
    suppliers_filename = fields.Char()

    sale_tax_id = fields.Many2one(
        'account.tax', string='Sales Tax (inclusive)',
        domain="[('type_tax_use','=','sale'),('company_id','=',company_id)]",
        help='Optional. Prices are treated as tax-inclusive; choose a '
             'price-included tax to break out VAT while keeping totals equal.')
    purchase_tax_id = fields.Many2one(
        'account.tax', string='Purchase Tax (inclusive)',
        domain="[('type_tax_use','=','purchase'),('company_id','=',company_id)]")

    bank_journal_id = fields.Many2one(
        'account.journal', string='Payments Journal',
        domain="[('type','in',('bank','cash')),('company_id','=',company_id)]")
    misc_journal_id = fields.Many2one(
        'account.journal', string='Opening Journal',
        domain="[('type','=','general'),('company_id','=',company_id)]")

    batch_ref = fields.Char(
        string='Batch Reference', required=True,
        default=lambda self: 'PETRO-IMP-%s' % fields.Datetime.now().strftime('%Y%m%d-%H%M%S'))
    clean_existing = fields.Boolean(
        string='Delete previous imports first', default=True,
        help='Remove any journal entries previously created by this importer '
             '(any batch) before importing, so the wizard is safely repeatable.')

    state = fields.Selection([
        ('draft', 'Draft'), ('queued', 'Running'), ('done', 'Done'),
    ], default='draft')
    job_id = fields.Many2one('petroleum.data.import.job', readonly=True)
    result_html = fields.Html(string='Result', readonly=True)

    # ------------------------------------------------------------------
    # Master-data helpers
    # ------------------------------------------------------------------
    def _get_opening_account(self):
        acc = self.env['account.account'].search([
            ('company_ids', 'in', self.company_id.id),
            ('name', '=', 'Opening Balance Import'),
        ], limit=1)
        if not acc:
            acc = self.env['account.account'].create({
                'name': 'Opening Balance Import',
                'code': self._next_account_code('OPNBAL'),
                'account_type': 'equity',
                'company_ids': [Command.link(self.company_id.id)],
            })
        return acc

    def _next_account_code(self, base):
        existing = self.env['account.account'].search_count([
            ('company_ids', 'in', self.company_id.id), ('code', '=', base)])
        return base if not existing else '%s%s' % (base, existing)

    def _ensure_bank_journal(self, name, code):
        journal = self.env['account.journal'].search([
            ('type', '=', 'bank'), ('code', '=', code),
            ('company_id', '=', self.company_id.id)], limit=1)
        if journal:
            return journal
        account = self.env['account.account'].search([
            ('company_ids', 'in', self.company_id.id),
            ('name', '=', 'Bank - %s' % name)], limit=1)
        if not account:
            account = self.env['account.account'].create({
                'name': 'Bank - %s' % name,
                'code': self._next_account_code('BNK%s' % code),
                'account_type': 'asset_cash',
                'company_ids': [Command.link(self.company_id.id)],
            })
        return self.env['account.journal'].create({
            'name': name,
            'code': code,
            'type': 'bank',
            'company_id': self.company_id.id,
            'default_account_id': account.id,
        })

    def _build_bank_map(self):
        bank_map = {}
        for payer, (jname, jcode) in BANK_NAME_MAP.items():
            bank_map[payer] = self._ensure_bank_journal(jname, jcode)
        fallback = self.bank_journal_id or self._ensure_bank_journal(
            'Bank - Unspecified', 'BNKU')
        return bank_map, fallback

    def _receivable_account(self, partner):
        acc = partner.with_company(self.company_id).property_account_receivable_id
        if not acc:
            acc = self.env['account.account'].search([
                ('company_ids', 'in', self.company_id.id),
                ('account_type', '=', 'asset_receivable')], limit=1)
        return acc

    def _payable_account(self, partner):
        acc = partner.with_company(self.company_id).property_account_payable_id
        if not acc:
            acc = self.env['account.account'].search([
                ('company_ids', 'in', self.company_id.id),
                ('account_type', '=', 'liability_payable')], limit=1)
        return acc

    def _get_partner(self, st, name, is_customer=False, is_supplier=False):
        key = name.strip().upper()
        cache = st['partner_cache']
        partner = cache.get(key)
        if not partner:
            partner = self.env['res.partner'].search(
                [('name', '=ilike', name.strip())], limit=1)
            if not partner:
                partner = self.env['res.partner'].with_context(**IMPORT_CTX).create({
                    'name': name.strip(),
                    'company_type': 'company',
                })
            cache[key] = partner
        vals = {}
        if is_customer and partner.customer_rank < 1:
            vals['customer_rank'] = 1
        if is_supplier and partner.supplier_rank < 1:
            vals['supplier_rank'] = 1
        if vals:
            partner.write(vals)
        return partner

    def _get_fuel_product(self, st, grade):
        cache = st['product_cache']
        prod = cache.get(grade)
        if prod:
            return prod
        prod = self.env['product.product'].search([('default_code', '=', grade)], limit=1)
        if not prod:
            litre = self.env.ref('uom.product_uom_litre')
            income = st['sale_journal'].default_account_id
            expense = st['purchase_journal'].default_account_id
            tmpl_vals = {
                'name': {'PMS': 'PMS (Petrol)', 'AGO': 'AGO (Diesel)',
                         'IK': 'IK (Kerosene)'}.get(grade, grade),
                'default_code': grade,
                'type': 'consu',
                'uom_id': litre.id,
                'sale_ok': True,
                'purchase_ok': True,
                'fuel_ok': True,
                'taxes_id': [Command.set(st['sale_tax'].ids)] if st['sale_tax'] else False,
                'supplier_taxes_id': [Command.set(st['purchase_tax'].ids)] if st['purchase_tax'] else False,
            }
            if income:
                tmpl_vals['property_account_income_id'] = income.id
            if expense:
                tmpl_vals['property_account_expense_id'] = expense.id
            prod = self.env['product.product'].create(tmpl_vals)
        cache[grade] = prod
        return prod

    # ------------------------------------------------------------------
    # Workbook parsing (validated against the source files)
    # ------------------------------------------------------------------
    @staticmethod
    def _is_header(cells):
        return 'BALANCE' in cells and ('DEBIT' in cells or 'CREDIT' in cells)

    @staticmethod
    def _build_colmap(header):
        cm, grade = {}, None
        for idx, h in enumerate(header):
            if h in ('LOADING', 'LOADING DATE', 'DATE'):
                cm.setdefault('date', idx)
            elif h == 'LOADING POINT':
                cm['lp'] = idx
            elif h in ('TRUCKS', 'TRUCK'):
                cm['truck'] = idx
            elif h in GRADES:
                cm[h.lower()] = idx
                grade = h.lower()
            elif h == 'PRICE' and grade:
                cm[grade + '_price'] = idx
                grade = None
            elif h == 'SELLING PRICE':
                cm['sp'] = idx
            elif h == 'INVOICE NO':
                cm['inv'] = idx
            elif h == 'TRANS DATE':
                cm['tdate'] = idx
            elif h == 'DEBIT':
                cm['debit'] = idx
            elif h == 'CREDIT':
                cm['credit'] = idx
            elif h == 'BALANCE':
                cm['balance'] = idx
        return cm

    def _parse_sheet(self, ws, default_side):
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=18, values_only=True))
        up = [[str(c).strip().upper() if c is not None else '' for c in r] for r in rows]
        header_idxs = [i for i, c in enumerate(up) if self._is_header(c)]
        sections = []
        for si, hi in enumerate(header_idxs):
            end = header_idxs[si + 1] if si + 1 < len(header_idxs) else len(rows)
            cm = self._build_colmap(up[hi])
            if 'balance' not in cm:
                continue
            side = default_side
            for j in range(max(0, hi - 3), hi + 1):
                blob = ' '.join(up[j])
                if 'PAYABLE' in blob:
                    side = 'ap'
                elif 'RECEIVABLE' in blob:
                    side = 'ar'
            bf, txns, last_balance = 0.0, [], None
            for r in rows[hi + 1:end]:
                def get(k):
                    i = cm.get(k)
                    return r[i] if (i is not None and i < len(r)) else None
                bal = get('balance')
                if isinstance(bal, (int, float)):
                    last_balance = bal
                sp = get('sp')
                if isinstance(sp, str) and sp.strip().upper() == 'B/F':
                    bf = _num(bal)
                    continue
                d = _parse_date(get('date')) or _parse_date(get('tdate'))
                if d is None:
                    continue
                debit, credit = _num(get('debit')), _num(get('credit'))
                if debit == 0 and credit == 0:
                    continue
                lp = (str(get('lp')).strip().upper() if get('lp') else '')
                effect = (debit - credit) if side == 'ar' else (credit - debit)
                lines = []
                for g in ('pms', 'ago', 'ik'):
                    q, p = _num(get(g)), _num(get(g + '_price'))
                    if q and p:
                        lines.append((g.upper(), q, p))
                if lp in ('PAYMENT', 'REFUND') or (not lines and _num(sp) == 0):
                    kind = 'payment'
                else:
                    kind = 'loading'
                txns.append({
                    'date': d, 'effect': effect, 'debit': debit, 'credit': credit,
                    'kind': kind, 'lp': lp, 'lines': lines, 'sp': _num(sp),
                    'truck': str(get('truck') or ''),
                    'inv': str(get('inv')) if get('inv') not in (None, '') else '',
                })
            sections.append({'side': side, 'bf': bf, 'txns': txns,
                             'last_balance': last_balance or 0.0})
        return sections

    @staticmethod
    def _partner_name(ws, sn):
        title = None
        for r in ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True):
            title = r[0]
        name = (str(title).strip() if title else sn).strip()
        for suf in ('(PAYABLE)', '(RECEIVABLE)'):
            name = name.replace(suf, '').replace(suf.lower(), '').strip()
        return name or sn

    def _iter_partner_sections(self, binary, default_side):
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(binary)), data_only=True)
        try:
            for sn in wb.sheetnames:
                if sn in ('Sheet1', 'Sheet2'):
                    continue
                ws = wb[sn]
                name = self._partner_name(ws, sn)
                if name.upper() in CONTROL_NAMES or sn.upper() in CONTROL_NAMES:
                    continue
                secs = self._parse_sheet(ws, default_side)
                if secs:
                    yield name, secs
        finally:
            wb.close()

    @staticmethod
    def _serialize_txn(txn):
        row = dict(txn)
        row['date'] = txn['date'].isoformat()
        row['lines'] = list(txn.get('lines') or [])
        return row

    def _build_work_queue(self, sources):
        queue = []
        for binary, default_side in sources:
            for name, sections in self._iter_partner_sections(binary, default_side):
                for sec in sections:
                    queue.append({
                        'name': name,
                        'side': sec['side'],
                        'section': {
                            'side': sec['side'],
                            'bf': sec['bf'],
                            'last_balance': sec['last_balance'],
                            'txns': [self._serialize_txn(t) for t in sec['txns']],
                        },
                    })
        return queue

    @staticmethod
    def _deserialize_txns(txns):
        rows = []
        for txn in txns:
            row = dict(txn)
            row['date'] = fields.Date.to_date(txn['date'])
            rows.append(row)
        return rows

    # ------------------------------------------------------------------
    # Posting
    # ------------------------------------------------------------------
    def _post_opening(self, st, partner, side, amount, account, date):
        if self.company_id.currency_id.is_zero(amount):
            return
        if side == 'ar':
            party = {'debit': amount, 'credit': 0.0} if amount > 0 else {'debit': 0.0, 'credit': -amount}
        else:
            party = {'credit': amount, 'debit': 0.0} if amount > 0 else {'credit': 0.0, 'debit': -amount}
        counter = {'debit': party['credit'], 'credit': party['debit']}
        move = self.env['account.move'].with_context(**IMPORT_CTX).create({
            'move_type': 'entry',
            'journal_id': st['misc_journal'].id,
            'date': date,
            'ref': _('Opening balance %s') % partner.name,
            'company_id': st['company_id'],
            'petro_import_batch': st['batch_ref'],
            'line_ids': [
                Command.create({'account_id': account.id, 'partner_id': partner.id,
                                'name': _('Opening balance'), **party}),
                Command.create({'account_id': st['opening_acc'].id,
                                'name': _('Opening balance'), **counter}),
            ],
        })
        move.action_post()
        st['counters']['opening'] += 1

    def _post_loading(self, st, partner, side, txn):
        move_type = 'out_invoice' if side == 'ar' else 'in_invoice'
        journal = st['sale_journal'] if side == 'ar' else st['purchase_journal']
        tax = st['sale_tax'] if side == 'ar' else st['purchase_tax']
        line_cmds = []
        for grade, qty, price in txn['lines']:
            line_cmds.append(Command.create({
                'product_id': self._get_fuel_product(st, grade).id,
                'quantity': qty,
                'price_unit': price,
                'tax_ids': [Command.set(tax.ids)] if tax else [Command.clear()],
            }))
        if not line_cmds:
            line_cmds.append(Command.create({
                'product_id': self._get_fuel_product(st, 'AGO').id,
                'name': _('Fuel (unspecified)'),
                'quantity': 1.0,
                'price_unit': txn['sp'],
                'tax_ids': [Command.set(tax.ids)] if tax else [Command.clear()],
            }))
        narration = ' '.join(filter(None, [txn['lp'], txn['truck']]))
        move = self.env['account.move'].with_context(**IMPORT_CTX).create({
            'move_type': move_type,
            'journal_id': journal.id,
            'partner_id': partner.id,
            'invoice_date': txn['date'],
            'date': txn['date'],
            'ref': txn['inv'] or narration or False,
            'narration': narration,
            'company_id': st['company_id'],
            'petro_import_batch': st['batch_ref'],
            'invoice_line_ids': line_cmds,
        })
        move.action_post()
        st['counters']['invoice' if side == 'ar' else 'bill'] += 1

    def _post_payment(self, st, partner, side, txn, account):
        effect = txn['effect']
        if self.company_id.currency_id.is_zero(effect):
            return
        if side == 'ar':
            party = {'debit': effect, 'credit': 0.0} if effect > 0 else {'debit': 0.0, 'credit': -effect}
        else:
            party = {'credit': effect, 'debit': 0.0} if effect > 0 else {'credit': 0.0, 'debit': -effect}
        bank = {'debit': party['credit'], 'credit': party['debit']}
        payer = (txn['truck'] or '').strip().upper()
        journal = st['bank_map'].get(payer, st['bank_fallback'])
        label = ' '.join(filter(None, [txn['lp'] or _('Payment'), txn['truck']]))
        move = self.env['account.move'].with_context(**IMPORT_CTX).create({
            'move_type': 'entry',
            'journal_id': journal.id,
            'date': txn['date'],
            'ref': label,
            'company_id': st['company_id'],
            'petro_import_batch': st['batch_ref'],
            'line_ids': [
                Command.create({'account_id': account.id, 'partner_id': partner.id,
                                'name': label, **party}),
                Command.create({'account_id': journal.default_account_id.id,
                                'partner_id': partner.id, 'name': label, **bank}),
            ],
        })
        move.action_post()
        st['counters']['payment'] += 1

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------
    def _init_st(self):
        """Build shared posting state (journals, caches, counters)."""
        company = self.company_id
        self = self.with_company(company)

        def _journal(domain, required_name):
            jrec = self.env['account.journal'].search(
                domain + [('company_id', '=', company.id)], limit=1)
            if not jrec:
                raise UserError(_('No %s journal found for company %s.')
                                % (required_name, company.name))
            return jrec

        misc_journal = self.misc_journal_id or _journal(
            [('type', '=', 'general')], 'Miscellaneous')
        bank_map, bank_fallback = self._build_bank_map()
        if not bank_fallback.default_account_id:
            raise UserError(_('The fallback bank journal "%s" has no default account.')
                            % bank_fallback.name)

        return {
            'company_id': company.id,
            'batch_ref': self.batch_ref,
            'partner_cache': {},
            'product_cache': {},
            'counters': {'opening': 0, 'invoice': 0, 'bill': 0, 'payment': 0},
            'sale_journal': _journal([('type', '=', 'sale')], 'Sales'),
            'purchase_journal': _journal([('type', '=', 'purchase')], 'Purchase'),
            'misc_journal': misc_journal,
            'bank_map': bank_map,
            'bank_fallback': bank_fallback,
            'sale_tax': self.sale_tax_id,
            'purchase_tax': self.purchase_tax_id,
            'opening_acc': self._get_opening_account(),
        }

    def _process_one_section(self, st, item, opening_date, cutoff, recon):
        name = item['name']
        sec = dict(item['section'])
        sec['txns'] = self._deserialize_txns(sec['txns'])
        side = sec['side']
        partner = self._get_partner(
            st, name, is_customer=(side == 'ar'), is_supplier=(side == 'ap'))
        account = (self._receivable_account(partner) if side == 'ar'
                   else self._payable_account(partner))
        if not account:
            raise UserError(_('No %s account for partner %s.') % (side, partner.name))
        pre = sum(t['effect'] for t in sec['txns'] if t['date'] < cutoff)
        opening = sec['bf'] + pre
        final = sec['bf'] + sum(t['effect'] for t in sec['txns'])
        self._post_opening(st, partner, side, opening, account, opening_date)
        for txn in sorted(
                [t for t in sec['txns'] if t['date'] >= cutoff],
                key=lambda t: t['date']):
            if txn['kind'] == 'loading':
                self._post_loading(st, partner, side, txn)
            else:
                self._post_payment(st, partner, side, txn, account)
        recon.append({
            'name': name,
            'side': side,
            'wb_final': final,
            'partner_id': partner.id,
            'account_id': account.id,
        })

    def action_import(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_('The python library "openpyxl" is required.'))
        if not self.customers_file and not self.suppliers_file:
            raise UserError(_('Upload at least one workbook.'))

        if self.clean_existing:
            self.with_context(**IMPORT_CTX)._delete_previous()

        sources = []
        if self.customers_file:
            sources.append((self.customers_file, 'ar'))
        if self.suppliers_file:
            sources.append((self.suppliers_file, 'ap'))

        queue = self._build_work_queue(sources)
        if not queue:
            raise UserError(_('No ledger sections found in the workbook(s).'))

        opening_date = fields.Date.to_date(self.cutoff_date) - datetime.timedelta(days=1)
        job = self.env['petroleum.data.import.job'].create({
            'name': _('Ledger import %s') % self.batch_ref,
            'company_id': self.company_id.id,
            'batch_ref': self.batch_ref,
            'cutoff_date': self.cutoff_date,
            'opening_date': opening_date,
            'sale_tax_id': self.sale_tax_id.id,
            'purchase_tax_id': self.purchase_tax_id.id,
            'bank_journal_id': self.bank_journal_id.id,
            'misc_journal_id': self.misc_journal_id.id,
            'queue_sections': queue,
            'total_sections': len(queue),
        })
        job._kick()

        eta = max(1, len(queue) // SECTION_BATCH + 1)
        html = (
            "<div><h3>Ledger import queued</h3>"
            "<p>Parsed <b>%d</b> partner ledger section(s) from the workbook(s).</p>"
            "<p>Posting runs in the background (about <b>%d</b> cron cycle(s)). "
            "Invoices, bills and payments will appear over the next few minutes.</p>"
            "<p>Batch reference: <b>%s</b></p>"
            "<p>When finished, run <b>Trading Desk → Configuration → "
            "Reconcile Imported Ledgers</b>.</p></div>"
        ) % (len(queue), eta, self.batch_ref)

        self.write({'state': 'queued', 'job_id': job.id, 'result_html': html})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def _delete_previous(self):
        moves = self.env['account.move'].search([
            ('petro_import_batch', '!=', False),
            ('company_id', '=', self.company_id.id)])
        if moves:
            moves.filtered(lambda m: m.state == 'posted').button_draft()
            moves.with_context(force_delete=True).unlink()

    def _odoo_balance(self, partner, account, side):
        self.env.cr.execute("""
            SELECT COALESCE(SUM(debit - credit), 0)
            FROM account_move_line aml
            JOIN account_move am ON am.id = aml.move_id
            WHERE aml.partner_id = %s AND aml.account_id = %s
              AND am.state = 'posted' AND am.company_id = %s
        """, (partner.id, account.id, self.company_id.id))
        bal = self.env.cr.fetchone()[0] or 0.0
        return bal if side == 'ar' else -bal

    def _build_report(self, recon, counters):
        cur = self.company_id.currency_id
        # Aggregate by partner+account+side: several worksheets (e.g. GAPCO /
        # GAPCO IRO) or multiple sections can map to one partner ledger.
        grouped = {}
        for name, side, wb_final, partner, account in recon:
            key = (partner.id, account.id, side)
            g = grouped.setdefault(key, {'names': set(), 'wb': 0.0,
                                         'partner': partner, 'account': account, 'side': side})
            g['names'].add(name)
            g['wb'] += wb_final
        rows, mismatches = [], 0
        tot_wb = tot_odoo = 0.0
        for g in grouped.values():
            odoo_bal = self._odoo_balance(g['partner'], g['account'], g['side'])
            wb_final = g['wb']
            diff = odoo_bal - wb_final
            tot_wb += wb_final
            tot_odoo += odoo_bal
            ok = cur.is_zero(diff)
            if not ok:
                mismatches += 1
            name = ' / '.join(sorted(g['names']))
            if not cur.is_zero(wb_final) or not ok:
                rows.append((name, g['side'], wb_final, odoo_bal, diff, ok))
        rows.sort(key=lambda r: r[0])
        c = counters

        def money(v):
            return '{:,.0f}'.format(v)

        head = (
            "<div>"
            "<h3>Petroleum import &mdash; batch %s</h3>"
            "<p>Created: <b>%d</b> opening entries, <b>%d</b> customer invoices, "
            "<b>%d</b> vendor bills, <b>%d</b> payments.</p>"
            "<p>Partners/sections reconciled: <b>%d</b> &mdash; "
            "<b style='color:%s'>%d mismatch(es)</b>.</p>"
            "<p>Workbook total: <b>%s</b> &nbsp; Odoo total: <b>%s</b> &nbsp; Diff: <b>%s</b></p>"
        ) % (self.batch_ref, c['opening'], c['invoice'], c['bill'], c['payment'],
             len(recon), 'red' if mismatches else 'green', mismatches,
             money(tot_wb), money(tot_odoo), money(tot_odoo - tot_wb))

        trs = []
        for name, side, wb_final, odoo_bal, diff, ok in rows:
            color = '' if ok else "background:#ffd5d5;"
            trs.append(
                "<tr style='%s'><td>%s</td><td>%s</td>"
                "<td style='text-align:right'>%s</td>"
                "<td style='text-align:right'>%s</td>"
                "<td style='text-align:right'>%s</td>"
                "<td style='text-align:center'>%s</td></tr>" % (
                    color, name, side.upper(), money(wb_final), money(odoo_bal),
                    money(diff), 'OK' if ok else 'CHECK'))
        table = (
            "<table class='table table-sm' style='width:100%%'>"
            "<thead><tr><th>Partner</th><th>Side</th>"
            "<th style='text-align:right'>Workbook</th>"
            "<th style='text-align:right'>Odoo</th>"
            "<th style='text-align:right'>Diff</th><th>Status</th></tr></thead>"
            "<tbody>%s</tbody></table></div>" % ''.join(trs))
        return head + table
