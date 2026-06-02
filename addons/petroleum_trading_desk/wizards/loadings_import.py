import base64
import datetime
import io
import logging
import re

from odoo import Command, _, fields, models
from odoo.exceptions import UserError

from odoo.addons.petroleum_trading_desk.models.loadings_import_job import (
    CONFIRM_BATCH,
    IMPORT_CTX,
)

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

GRADES = (
    ('pms', 11, 2, 3),
    ('ago', 12, 4, 5),
    ('ik', 13, 6, 7),
)
NOT_SOLD_NAMES = frozenset({'NOT SOLD', 'NOTSOLD', 'UNSOLD'})
CREATE_BATCH = 25


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def _parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return datetime.date.fromisoformat(s)
        except ValueError:
            pass
        for sep in ('/', '-', '.'):
            if sep in s:
                parts = s.split(sep)
                if len(parts) == 3:
                    try:
                        a, b, c = (int(p) for p in parts)
                    except ValueError:
                        return None
                    if a > 31:
                        y, m, d = a, b, c
                    else:
                        d, m, y = a, b, c
                    if y < 100:
                        y += 2000
                    try:
                        return datetime.date(y, m, d)
                    except ValueError:
                        return None
    return None


class PetroleumLoadingsImport(models.TransientModel):
    _name = 'petroleum.loadings.import'
    _description = 'Import Loadings Workbook'

    company_id = fields.Many2one(
        'res.company', string='Company', required=True,
        default=lambda self: self.env.company)
    loadings_file = fields.Binary(string='Loadings Workbook (.xlsx)', required=True)
    loadings_filename = fields.Char()
    not_sold_partner_id = fields.Many2one(
        'res.partner', string='Placeholder for "NOT SOLD"',
        domain="[('customer_rank', '>', 0)]",
        help='Loads marked NOT SOLD in the workbook are assigned to this '
             'customer and left in Draft.')
    skip_existing = fields.Boolean(
        string='Skip rows already imported', default=True,
        help='Skip a row when a deal already exists with the same date, truck, '
             'customer, supplier and product quantities.')
    auto_confirm = fields.Boolean(
        string='Confirm deals (background)', default=True,
        help='Queue sale order / purchase order / trip creation in a '
             'background job so large workbooks do not hit the HTTP timeout.')
    auto_load = fields.Boolean(
        string='Load and invoice (background)', default=False,
        help='Post customer invoices and vendor bills after confirming. '
             'Leave off when accounting was already imported from the '
             'customer/supplier ledger workbooks.')
    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft')
    job_id = fields.Many2one('petroleum.loadings.import.job', readonly=True)
    result_html = fields.Html(string='Result', readonly=True)

    # ------------------------------------------------------------------
    # Parsing (MAY LOADING REPORT layout: two header rows, data from row 3)
    # ------------------------------------------------------------------
    @staticmethod
    def _validate_header_rows(header1, header2):
        a1 = (header1[0] if header1 else '') or ''
        a2 = (header2[0] if header2 else '') or ''
        if str(a1).strip().upper() != 'LOADING':
            raise UserError(_(
                'Unrecognised workbook layout: cell A1 should be "LOADING" '
                '(found "%s").') % a1)
        if str(a2).strip().upper() != 'DATE':
            raise UserError(_(
                'Unrecognised workbook layout: cell A2 should be "DATE" '
                '(found "%s").') % a2)

    @staticmethod
    def _row_val(row, col, default=None):
        """Return a 1-based column value from an iter_rows tuple."""
        idx = col - 1
        if not row or idx >= len(row):
            return default
        return row[idx]

    def _parse_rows(self, ws):
        """Single-pass parse — never call ws.cell() on a read-only sheet."""
        sheet_rows = list(ws.iter_rows(min_row=1, max_col=17, values_only=True))
        if len(sheet_rows) < 2:
            raise UserError(_('Workbook sheet is empty.'))
        self._validate_header_rows(sheet_rows[0], sheet_rows[1])

        rows = []
        for row_idx, row in enumerate(sheet_rows[2:], start=3):
            date = _parse_date(self._row_val(row, 1))
            if not date:
                continue
            supplier_raw = self._row_val(row, 8)
            customer_raw = self._row_val(row, 9)
            vehicle = self._row_val(row, 10)
            if not supplier_raw and not customer_raw and not vehicle:
                continue
            lines = []
            for grade, qty_col, buy_col, sell_col in GRADES:
                qty = _num(self._row_val(row, qty_col))
                if not qty:
                    continue
                buy = self._row_val(row, buy_col)
                sell = self._row_val(row, sell_col)
                if buy is None and sell is None:
                    continue
                lines.append({
                    'grade': grade.upper(),
                    'quantity': qty,
                    'buy_price': _num(buy) if buy is not None else 0.0,
                    'sell_price': _num(sell) if sell is not None else _num(buy),
                })
            if not lines:
                continue
            rows.append({
                'row': row_idx,
                'date': date,
                'supplier_raw': str(supplier_raw or '').strip(),
                'customer_raw': str(customer_raw or '').strip(),
                'vehicle': str(vehicle or '').strip(),
                'lines': lines,
                'total_qty': sum(l['quantity'] for l in lines),
            })
        return rows

    # ------------------------------------------------------------------
    # Master data
    # ------------------------------------------------------------------
    def _ensure_not_sold_partner(self):
        if self.not_sold_partner_id:
            return self.not_sold_partner_id
        partner = self.env['res.partner'].search(
            [('name', '=ilike', 'NOT SOLD')], limit=1)
        if not partner:
            partner = self.env['res.partner'].with_context(**IMPORT_CTX).create({
                'name': 'NOT SOLD',
                'company_type': 'company',
                'customer_rank': 1,
            })
        return partner

    def _prepare_fuel_products_once(self, st):
        """Fuel is brokered back-to-back — avoid stock moves during confirm."""
        for grade in ('PMS', 'AGO', 'IK'):
            self._get_fuel_product(st, grade)
        products = self.env['product.product'].concat(*st['product_cache'].values())
        templates = products.product_tmpl_id
        vals = {}
        if any(t.is_storable for t in templates):
            vals['is_storable'] = False
        if any(t.invoice_policy != 'order' for t in templates):
            vals['invoice_policy'] = 'order'
        if 'purchase_method' in templates._fields and any(
                t.purchase_method != 'purchase' for t in templates):
            vals['purchase_method'] = 'purchase'
        if vals:
            templates.with_context(**IMPORT_CTX).write(vals)

    @staticmethod
    def _split_supplier(name):
        no_invoice = bool(re.search(r'\bNO\s+INVOICE\b', name, re.I))
        clean = re.sub(r'\s+NO\s+INVOICE\s*$', '', name, flags=re.I).strip()
        return clean or name, no_invoice

    def _get_partner(self, st, name, is_customer=False, is_supplier=False):
        key = name.strip().upper()
        cache = st['partner_cache']
        partner = cache.get(key)
        if not partner:
            partner = self.env['res.partner'].search(
                [('name', '=ilike', name.strip())], limit=1)
            if not partner:
                partner = self.env['res.partner'].with_context(**IMPORT_CTX).create({
                    'name': name.strip(),
                    'company_type': 'company',
                })
            cache[key] = partner
        vals = {}
        if is_customer and partner.customer_rank < 1:
            vals['customer_rank'] = 1
        if is_supplier and partner.supplier_rank < 1:
            vals['supplier_rank'] = 1
        if vals:
            partner.with_context(**IMPORT_CTX).write(vals)
        return partner

    def _get_truck(self, st, plate):
        if not plate:
            return self.env['truck.management']
        key = plate.strip().upper()
        cache = st['truck_cache']
        truck = cache.get(key)
        if not truck:
            truck = self.env['truck.management'].search(
                [('name', '=ilike', plate.strip())], limit=1)
            if not truck:
                truck = self.env['truck.management'].with_context(**IMPORT_CTX).create({
                    'name': plate.strip().upper(),
                    'capacity': 36000,
                })
            cache[key] = truck
        return truck

    def _get_fuel_product(self, st, grade):
        cache = st['product_cache']
        prod = cache.get(grade)
        if prod:
            return prod
        prod = self.env['product.product'].search(
            [('default_code', '=', grade)], limit=1)
        if not prod:
            litre = self.env.ref('uom.product_uom_litre')
            prod = self.env['product.product'].with_context(**IMPORT_CTX).create({
                'name': {'PMS': 'PMS (Petrol)', 'AGO': 'AGO (Diesel)',
                         'IK': 'IK (Kerosene)'}.get(grade, grade),
                'default_code': grade,
                'type': 'consu',
                'uom_id': litre.id,
                'sale_ok': True,
                'purchase_ok': True,
                'fuel_ok': True,
            })
        cache[grade] = prod
        return prod

    @staticmethod
    def _row_key(row, partner, supplier, truck):
        return (
            row['date'],
            partner.id,
            truck.id if truck else 0,
            supplier.id,
            round(row['total_qty'], 3),
        )

    def _load_existing_keys(self, parsed):
        if not parsed:
            return set()
        dates = [r['date'] for r in parsed]
        deals = self.env['petroleum.deal'].search([
            ('date', '>=', min(dates)),
            ('date', '<=', max(dates)),
            ('state', '!=', 'cancel'),
        ])
        keys = set()
        for deal in deals:
            supplier = deal.line_ids[:1].supplier_id.id if deal.line_ids else 0
            keys.add((
                deal.date,
                deal.partner_id.id,
                deal.truck_id.id if deal.truck_id else 0,
                supplier,
                round(deal.total_qty, 3),
            ))
        return keys

    def _build_deal_vals(self, st, row, partner, supplier, truck, not_sold):
        line_cmds = []
        for line in row['lines']:
            product = self._get_fuel_product(st, line['grade'])
            line_cmds.append(Command.create({
                'product_id': product.id,
                'quantity': line['quantity'],
                'buy_price': line['buy_price'],
                'sell_price': line['sell_price'],
                'supplier_id': supplier.id,
            }))
        notes_parts = [_('Imported from row %s.') % row['row']]
        if row.get('supplier_no_invoice'):
            notes_parts.append(_('Supplier bill not yet received (NO INVOICE).'))
        if not_sold:
            notes_parts.append(_('Customer not assigned at loading time.'))
        return {
            'date': row['date'],
            'partner_id': partner.id,
            'truck_id': truck.id if truck else False,
            'company_id': self.company_id.id,
            'line_ids': line_cmds,
            'notes': ' '.join(notes_parts),
        }

    # ------------------------------------------------------------------
    # Main
    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_('The python library "openpyxl" is required.'))
        if not self.loadings_file:
            raise UserError(_('Upload a loadings workbook.'))

        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(self.loadings_file)),
            data_only=True)
        ws = wb.active
        parsed = self._parse_rows(ws)
        wb.close()

        if not parsed:
            raise UserError(_('No loading rows found in the workbook.'))

        not_sold_partner = self._ensure_not_sold_partner()
        st = {
            'partner_cache': {},
            'truck_cache': {},
            'product_cache': {},
        }
        self._prepare_fuel_products_once(st)

        existing_keys = self._load_existing_keys(parsed) if self.skip_existing else set()
        Deal = self.env['petroleum.deal'].with_company(self.company_id).with_context(**IMPORT_CTX)

        counters = {
            'created': 0, 'skipped': 0, 'queued': 0,
            'draft': 0, 'errors': 0,
        }
        error_rows = []
        to_create = []

        for row in parsed:
            try:
                supplier_name, no_invoice = self._split_supplier(row['supplier_raw'])
                if not supplier_name:
                    raise UserError(_('Missing supplier.'))
                supplier = self._get_partner(st, supplier_name, is_supplier=True)

                customer_name = row['customer_raw'] or 'NOT SOLD'
                not_sold = customer_name.strip().upper() in NOT_SOLD_NAMES
                partner = not_sold_partner if not_sold else self._get_partner(
                    st, customer_name, is_customer=True)

                truck = self._get_truck(st, row['vehicle'])
                row['supplier_no_invoice'] = no_invoice

                key = self._row_key(row, partner, supplier, truck)
                if key in existing_keys:
                    counters['skipped'] += 1
                    continue

                vals = self._build_deal_vals(
                    st, row, partner, supplier, truck, not_sold)
                to_create.append(vals)
                existing_keys.add(key)

                if not_sold or not self.auto_confirm:
                    counters['draft'] += 1
                elif not truck:
                    counters['draft'] += 1
                    error_rows.append((row['row'], _('No truck — will stay in draft.')))
                else:
                    counters['queued'] += 1

            except UserError as exc:
                counters['errors'] += 1
                error_rows.append((row['row'], str(exc)))
            except Exception as exc:  # noqa: BLE001
                counters['errors'] += 1
                error_rows.append((row['row'], str(exc)))
                _logger.exception('Loadings import prepare row %s failed', row['row'])

        created_deals = self.env['petroleum.deal']
        for offset in range(0, len(to_create), CREATE_BATCH):
            batch_vals = to_create[offset:offset + CREATE_BATCH]
            created_deals |= Deal.create(batch_vals)
            counters['created'] += len(batch_vals)

        job_id = False
        if self.auto_confirm and counters['queued']:
            draft_deals = created_deals.filtered(lambda d: d.state == 'draft' and d.truck_id)
            not_sold_partner_id = not_sold_partner.id
            queue_ids = [
                d.id for d in draft_deals
                if d.partner_id.id != not_sold_partner_id
            ]
            if queue_ids:
                job = self.env['petroleum.loadings.import.job'].create({
                    'name': _('Loadings %s') % fields.Datetime.now().strftime('%Y-%m-%d %H:%M'),
                    'company_id': self.company_id.id,
                    'auto_load': self.auto_load,
                    'deal_ids': [(6, 0, queue_ids)],
                    'queue_deal_ids': queue_ids,
                    'total_count': len(queue_ids),
                })
                job_id = job.id
                job._kick()

        html = self._build_report(len(parsed), counters, error_rows)
        self.write({
            'state': 'done',
            'result_html': html,
            'job_id': job_id,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def _build_report(self, total_rows, counters, error_rows):
        c = counters

        def row_html(msg, danger=False):
            style = "background:#ffd5d5;" if danger else ''
            return "<tr style='%s'><td>%s</td></tr>" % (style, msg)

        head = (
            "<div><h3>Loadings import</h3>"
            "<p>Workbook rows parsed: <b>%d</b></p>"
            "<p>Created: <b>%d</b> &nbsp; Skipped (duplicate): <b>%d</b> &nbsp; "
            "Queued for confirmation: <b>%d</b> &nbsp; "
            "Left in draft: <b>%d</b> &nbsp; "
            "Errors: <b style='color:%s'>%d</b></p>"
        ) % (
            total_rows, c['created'], c['skipped'], c['queued'],
            c['draft'], 'red' if c['errors'] else 'green', c['errors'],
        )

        if c['queued']:
            head += (
                "<p><b>Confirmation is running in the background</b> "
                "(about %d deals per minute). Deals appear as Confirmed "
                "within a few minutes. Refresh the Deals list to check progress.</p>"
            ) % max(1, c['queued'] // CONFIRM_BATCH + 1)

        notes = (
            "<p><i>NOT SOLD</i> rows stay in draft on partner <b>%s</b>. "
            "Rows tagged <i>NO INVOICE</i> are confirmed but not auto-invoiced "
            "on the purchase side when Load and invoice is enabled.</p>"
        ) % (self._ensure_not_sold_partner().display_name)

        err_table = ''
        if error_rows:
            rows = ''.join(
                row_html(_('Row %s: %s') % (r, msg), danger=True)
                for r, msg in error_rows[:50])
            if len(error_rows) > 50:
                rows += row_html(_('… and %d more.') % (len(error_rows) - 50))
            err_table = (
                "<table class='table table-sm' style='width:100%%'>"
                "<thead><tr><th>Notes</th></tr></thead><tbody>%s</tbody></table>"
            ) % rows

        return head + notes + err_table + '</div>'
